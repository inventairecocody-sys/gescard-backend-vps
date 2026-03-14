"""
Générateur de rapport Excel GESCARD — Multi-niveaux hiérarchiques
Usage: python generer_rapport_excel.py '<json_data>'
Retourne le fichier en base64 sur stdout

Niveaux :
  direction    → Administrateur  : vue nationale (5 onglets)
  coordination → Gestionnaire    : sa coordination, agences, sites (4 onglets)
  agence       → Chef d'équipe   : son agence + sites (3 onglets)
  site         → Opérateur       : son site seul (2 onglets)
"""
import sys, json, base64, io, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─── Palette ─────────────────────────────────────────────────────────────────
C_ORANGE = 'F77F00'
C_BLUE   = '0077B6'
C_GREEN  = '16a34a'
C_RED    = 'dc2626'
C_TEAL   = '0d9488'
C_PURPLE = '6d28d9'
C_DARK   = '1A1A1A'
C_GRAY   = '6B7280'
C_WHITE  = 'FFFFFF'
C_LIGHT  = 'F9F9F9'
C_WARN   = 'FFF3E0'
C_SUCC   = 'F0FDF4'
C_FAIL   = 'FEF2F2'

NIVEAU_COLOR = {
    'direction':    C_ORANGE,
    'coordination': C_BLUE,
    'agence':       C_TEAL,
    'site':         C_PURPLE,
}

NIVEAU_LABEL = {
    'direction':    'DIRECTION CENTRALE — VUE NATIONALE',
    'coordination': 'COORDINATION',
    'agence':       'AGENCE',
    'site':         'SITE',
}

# ─── Helpers styles ──────────────────────────────────────────────────────────
def _side(c='CCCCCC'): return Side(style='thin', color=c)
def _border(c='CCCCCC'): return Border(left=_side(c), right=_side(c), top=_side(c), bottom=_side(c))
def _fill(c): return PatternFill('solid', start_color=c, end_color=c)
def _font(bold=False, color=C_DARK, size=10): return Font(name='Arial', bold=bold, color=color, size=size)
def _align(h='left', wrap=False): return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

def taux_color(t):
    if t >= 75: return C_GREEN
    if t >= 50: return C_ORANGE
    return C_RED

def taux_label(t):
    if t >= 75: return '🏆 Excellent'
    if t >= 50: return '📈 En progression'
    return '⚠️ À améliorer'

def fmt(n):  return f"{int(n):,}".replace(',', ' ')
def fmtT(t): return f"{t:.2f}".replace('.', ',') + '%'
def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def wcell(ws, row, col, val, bold=False, color=C_DARK, bg=C_WHITE,
          size=10, align='left', wrap=False, brd=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _font(bold, color, size)
    c.fill = _fill(bg)
    c.alignment = _align(align, wrap)
    if brd: c.border = _border()
    return c

def header_row(ws, row, headers, widths, bg=C_ORANGE):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _font(True, C_WHITE, 10)
        c.fill = _fill(bg)
        c.alignment = _align('center')
        c.border = _border(bg)
        cw(ws, i, w)
    ws.row_dimensions[row].height = 22

def data_row(ws, row, values, aligns=None, colors=None, bg=None, bold_cols=None):
    alt = bg if bg else (C_LIGHT if row % 2 == 0 else C_WHITE)
    aligns     = aligns or ['left'] * len(values)
    colors     = colors or [C_DARK] * len(values)
    bold_cols  = bold_cols or []
    for i, v in enumerate(values, 1):
        wcell(ws, row, i, v, bold=(i in bold_cols),
              color=colors[i-1], bg=alt, align=aligns[i-1])
    ws.row_dimensions[row].height = 18

def sheet_title(ws, text, subtitle, nb_cols, bg):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f'A1:{get_column_letter(nb_cols)}1')
    c = ws.cell(1, 1, text)
    c.font = _font(True, C_WHITE, 14)
    c.fill = _fill(bg)
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f'A2:{get_column_letter(nb_cols)}2')
    c2 = ws.cell(2, 1, subtitle)
    c2.font = _font(False, C_WHITE, 9)
    c2.fill = _fill(bg)
    c2.alignment = _align('center')
    ws.row_dimensions[2].height = 18

def section_title(ws, row, text, nb_cols, bg=C_WARN):
    ws.merge_cells(f'A{row}:{get_column_letter(nb_cols)}{row}')
    c = ws.cell(row, 1, text)
    c.font = _font(True, C_DARK, 11)
    c.fill = _fill(bg)
    c.alignment = _align('left')
    ws.row_dimensions[row].height = 26
    return row + 1

def totals_row(ws, row, nb_cols, values, bg='333333'):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = _font(True, C_WHITE, 10)
        c.fill = _fill(bg)
        c.alignment = _align('center')
        c.border = _border()
    ws.row_dimensions[row].height = 20

def analyse_block(ws, row, lignes, nb_cols):
    for ligne in lignes:
        ws.merge_cells(f'A{row}:{get_column_letter(nb_cols)}{row}')
        c = ws.cell(row, 1, ligne)
        if not ligne:
            ws.row_dimensions[row].height = 8
            row += 1
            continue
        if ligne.startswith('📊') or ligne.startswith('🔎') or ligne.startswith('📋'):
            c.font = _font(True, C_DARK, 10);  c.fill = _fill(C_WARN)
        elif ligne.startswith('🏆') or ligne.startswith('✅'):
            c.font = _font(False, C_GREEN, 10); c.fill = _fill(C_SUCC)
        elif ligne.startswith('⚠️') or ligne.startswith('🚨') or ligne.startswith('   →'):
            c.font = _font(False, C_RED, 10);   c.fill = _fill(C_FAIL)
        elif ligne.startswith('💡') or ligne.startswith('→'):
            c.font = _font(True, C_ORANGE, 10); c.fill = _fill(C_WARN)
        else:
            c.font = _font(False, C_DARK, 10);  c.fill = _fill(C_WHITE)
        c.alignment = _align('left', True)
        ws.row_dimensions[row].height = 18
        row += 1
    return row

# ─── Analyses contextuelles ──────────────────────────────────────────────────

def analyse_direction(data):
    coords   = data.get('coordinations', [])
    sites    = data.get('sites', [])
    moy_nat  = sum(c['tauxRetrait'] for c in coords) / len(coords) if coords else 0
    alertes  = [c for c in coords if c['tauxRetrait'] < 60]
    top      = sorted(coords, key=lambda c: c['tauxRetrait'], reverse=True)[:3]
    bas      = sorted(coords, key=lambda c: c['tauxRetrait'])[:3]
    restants = sum(s['restants'] for s in sites)

    lignes = [
        f"📊 ANALYSE NATIONALE — {len(coords)} coordinations  |  {fmt(restants)} cartes en attente de retrait par les requérants",
        f"Taux moyen national : {fmtT(moy_nat)}  |  Seuil d'alerte direction : 60%",
        "",
        "🏆 Top 3 coordinations : " + "  ·  ".join(f"{c['coordination']} ({fmtT(c['tauxRetrait'])})" for c in top),
        f"⚠️ Coordinations sous le seuil de 60% : {len(alertes)}",
    ]
    for c in bas:
        ecart = moy_nat - c['tauxRetrait']
        lignes.append(f"   → {c['coordination']} : {fmtT(c['tauxRetrait'])} — {ecart:.1f} pts sous la moyenne nationale")
    lignes.append("")
    if moy_nat >= 75:
        lignes += [
            "💡 RECOMMANDATION STRATÉGIQUE : Performance nationale excellente.",
            "→ Documenter les pratiques des coordinations leaders et les déployer nationalement.",
            "→ Engager la phase de clôture pour les requérants n'ayant pas retiré leur carte.",
            "→ Préparer le rapport final de distribution pour les autorités compétentes.",
        ]
    elif moy_nat >= 50:
        lignes += [
            "💡 RECOMMANDATION STRATÉGIQUE : Performance satisfaisante mais hétérogène.",
            f"→ Déclencher un audit opérationnel sur les {len(alertes)} coordination(s) sous 60%.",
            "→ Réallouer des ressources des zones performantes vers les zones critiques.",
            "→ Instaurer un reporting hebdomadaire national avec indicateurs par coordination.",
        ]
    else:
        lignes += [
            "🚨 SITUATION CRITIQUE — mobilisation nationale requise.",
            "→ Constituer une cellule de crise nationale avec réunion quotidienne.",
            "→ Déployer des équipes renforcées dans toutes les coordinations sous 50%.",
            "→ Lancer une campagne de sensibilisation pour les requérants non venus.",
            "→ Fixer des objectifs hebdomadaires contraignants avec reporting quotidien.",
        ]
    return lignes

def analyse_coordination(data, ctx):
    coord_nom  = ctx.get('coordination', 'votre coordination')
    agences    = data.get('agences', [])
    sites      = data.get('sites', [])
    coords_all = data.get('coordinations', [])
    moy_nat    = sum(c['tauxRetrait'] for c in coords_all) / len(coords_all) if coords_all else 0
    taux_coord = data.get('tauxRetrait', 0)
    ecart      = taux_coord - moy_nat
    signe      = f"+{ecart:.1f}" if ecart >= 0 else f"{ecart:.1f}"
    alertes_ag = [a for a in agences if a['taux_retrait'] < 55]
    restants   = sum(s['restants'] for s in sites)
    top_ag     = sorted(agences, key=lambda a: a['taux_retrait'], reverse=True)[:2]
    ag_crit    = max(agences, key=lambda a: a['cartes_restantes']) if agences else None

    lignes = [
        f"🔎 ANALYSE DE LA COORDINATION {coord_nom.upper()}",
        f"Taux : {fmtT(taux_coord)}  |  Moyenne nationale : {fmtT(moy_nat)}  |  Écart : {signe} pts",
        f"{fmt(restants)} cartes restantes — requérants n'ayant pas encore retiré leur carte",
        "",
    ]
    if top_ag:
        lignes.append("🏆 Meilleures agences : " + "  ·  ".join(f"{a['agence_nom']} ({fmtT(a['taux_retrait'])})" for a in top_ag))
    if ag_crit:
        pct = (ag_crit['cartes_restantes'] / restants * 100) if restants else 0
        lignes.append(f"⚠️ Agence concentrant le plus de retard : {ag_crit['agence_nom']} — {fmt(ag_crit['cartes_restantes'])} cartes ({pct:.0f}% du retard)")
    if alertes_ag:
        lignes.append(f"⚠️ Agences sous 55% ({len(alertes_ag)}) : " + ", ".join(a['agence_nom'] for a in alertes_ag))
    lignes += [
        "",
        f"💡 RECOMMANDATION : {'Performance au-dessus de la moyenne nationale.' if ecart >= 0 else 'Performance en-dessous de la moyenne nationale.'}",
        "→ Demander un rapport hebdomadaire aux chefs d'équipe des agences en retard.",
        "→ Fixer un objectif de taux cible par agence pour la prochaine période.",
        "→ Renforcer les sites où les requérants ne se présentent pas.",
    ]
    return lignes

def analyse_agence(data, ctx):
    agence_nom = ctx.get('agence', 'votre agence')
    sites      = data.get('sites', [])
    agences    = data.get('agences', [])
    taux_ag    = data.get('tauxRetrait', 0)
    restants   = sum(s['restants'] for s in sites)
    alertes_s  = [s for s in sites if s['tauxRetrait'] < 50]
    top_s      = sorted(sites, key=lambda s: s['tauxRetrait'], reverse=True)[:2]
    ags_sorted = sorted(agences, key=lambda a: a['taux_retrait'], reverse=True)
    rang       = next((i+1 for i, a in enumerate(ags_sorted) if a['agence_nom'] == agence_nom), None)
    rang_str   = f"classée {rang}e/{len(agences)} dans la coordination" if rang else ""

    lignes = [
        f"📋 ANALYSE DE L'AGENCE {agence_nom.upper()}",
        f"Taux : {fmtT(taux_ag)}  |  {rang_str}  |  {len(sites)} site(s)",
        f"{fmt(restants)} cartes restantes — requérants n'ayant pas encore retiré leur carte",
        "",
    ]
    if top_s:
        lignes.append("🏆 Sites les plus performants : " + "  ·  ".join(f"{s['site']} ({fmtT(s['tauxRetrait'])})" for s in top_s))
    if alertes_s:
        lignes.append(f"⚠️ Sites prioritaires (taux < 50%) : {len(alertes_s)} site(s)")
        for s in alertes_s[:4]:
            lignes.append(f"   → {s['site']} : {fmtT(s['tauxRetrait'])} — {fmt(s['restants'])} cartes en attente")
    lignes += [
        "",
        "💡 RECOMMANDATION OPÉRATIONNELLE :",
        f"→ Concentrer les efforts sur les {len(alertes_s)} site(s) en retard cette semaine.",
        "→ Vérifier que les requérants des sites critiques ont été contactés.",
        "→ Envisager des permanences supplémentaires sur les sites à fort volume restant.",
    ]
    return lignes

def analyse_site(data, ctx):
    site_nom = ctx.get('site', 'votre site')
    sites    = data.get('sites', [])
    taux     = data.get('tauxRetrait', 0)
    restants = data.get('restants', 0)
    retires  = data.get('retires', 0)
    s_sorted = sorted(sites, key=lambda s: s['tauxRetrait'], reverse=True)
    rang     = next((i+1 for i, s in enumerate(s_sorted) if s['site'] == site_nom), None)
    rang_str = f"classé {rang}e/{len(sites)} dans l'agence" if rang else ""

    rythme = ""
    if retires > 0 and restants > 0:
        jours = max(1, round(restants / max(retires / 30, 1)))
        rythme = f"Au rythme actuel, environ {jours} jour(s) pour traiter les cartes restantes."

    lignes = [
        f"📋 ÉTAT DU SITE {site_nom.upper()}",
        f"Taux : {fmtT(taux)}  |  {rang_str}",
        f"{fmt(restants)} requérant(s) n'ont pas encore retiré leur carte.",
    ]
    if rythme: lignes.append(rythme)
    lignes.append("")
    if taux >= 75:
        lignes += [
            "✅ Votre site affiche une performance excellente.",
            "→ Maintenir le rythme et signaler les requérants injoignables à votre chef d'équipe.",
        ]
    elif taux >= 50:
        lignes += [
            "💡 Performance satisfaisante — des efforts restent nécessaires.",
            f"→ Contacter les {fmt(restants)} requérants n'ayant pas retiré leur carte.",
            "→ Signaler les cas de requérants injoignables à votre chef d'équipe.",
        ]
    else:
        lignes += [
            "🚨 Taux insuffisant — action immédiate requise.",
            f"→ Établir la liste des {fmt(restants)} requérants et lancer les relances.",
            "→ Informer votre chef d'équipe et demander un renfort ou une permanence.",
            "→ Identifier les requérants avec problèmes de documents et les orienter.",
        ]
    return lignes

# ─── ONGLET RÉSUMÉ ────────────────────────────────────────────────────────────

def create_resume_sheet(wb, data, ctx, niveau):
    ws = wb.active
    ws.title = "📊 Résumé"
    bg        = NIVEAU_COLOR[niveau]
    nom       = ctx.get('coordination') or ctx.get('agence') or ctx.get('site') or ''
    titre_txt = f"RAPPORT D'ANALYSE GESCARD — {NIVEAU_LABEL[niveau]}" + (f" : {nom.upper()}" if nom else "")
    sous      = f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}  |  {ctx.get('nomUtilisateur','')}"

    sheet_title(ws, titre_txt, sous, 8, bg)

    # KPIs
    row = 4
    ws.merge_cells('A4:H4')
    ws.cell(4, 1, "INDICATEURS CLÉS").font = _font(True, C_DARK, 11)
    ws.row_dimensions[4].height = 26

    kpis = [
        ("Total cartes",     fmt(data['total']),         C_ORANGE),
        ("Cartes retirées",  fmt(data['retires']),        C_GREEN),
        ("Cartes restantes", fmt(data['restants']),       C_BLUE),
        ("Taux de retrait",  fmtT(data['tauxRetrait']),   taux_color(data['tauxRetrait'])),
    ]
    for i, (lbl, val, color) in enumerate(kpis):
        col = i * 2 + 1
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        c = ws.cell(5, col, val)
        c.font = _font(True, color, 22)
        c.alignment = _align('center')
        c.fill = _fill(C_LIGHT)
        c.border = Border(top=Side(style='thick', color=color), bottom=_side(), left=_side(), right=_side())
        ws.row_dimensions[5].height = 42
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
        c2 = ws.cell(6, col, lbl)
        c2.font = _font(False, C_GRAY, 9)
        c2.alignment = _align('center')
        c2.fill = _fill(C_LIGHT)
        ws.row_dimensions[6].height = 18

    for i in range(1, 9):
        cw(ws, i, 18)

    row = 8
    ws.merge_cells('A8:H8')
    ws.cell(8, 1, "ANALYSE ET RECOMMANDATIONS").font = _font(True, C_DARK, 11)
    ws.row_dimensions[8].height = 26

    fn = {
        'direction':    analyse_direction,
        'coordination': lambda d: analyse_coordination(d, ctx),
        'agence':       lambda d: analyse_agence(d, ctx),
        'site':         lambda d: analyse_site(d, ctx),
    }[niveau]
    analyse_block(ws, 9, fn(data), 8)

# ─── ONGLET COORDINATIONS (direction uniquement) ──────────────────────────────

def create_coordinations_sheet(wb, data, ctx):
    ws = wb.create_sheet("🏢 Par Coordination")
    coords  = sorted(data.get('coordinations', []), key=lambda c: c['tauxRetrait'], reverse=True)
    moy_nat = sum(c['tauxRetrait'] for c in coords) / len(coords) if coords else 0

    sheet_title(ws, "STATISTIQUES PAR COORDINATION", "Vue nationale — Direction centrale", 8, C_ORANGE)
    headers = ['Rang', 'Coordination', 'Total', 'Retirées', 'Restantes', 'Taux', 'Écart moy. nat.', 'Statut']
    widths  = [7, 28, 14, 14, 14, 12, 16, 18]
    header_row(ws, 3, headers, widths, C_ORANGE)

    for i, c in enumerate(coords):
        r     = i + 4
        ecart = c['tauxRetrait'] - moy_nat
        ecart_str = (f"+{ecart:.1f} pts" if ecart >= 0 else f"{ecart:.1f} pts")
        data_row(ws, r,
            [i+1, c['coordination'], fmt(c['total']), fmt(c['retires']),
             fmt(c['restants']), fmtT(c['tauxRetrait']), ecart_str, taux_label(c['tauxRetrait'])],
            aligns=['center','left','center','center','center','center','center','left'],
            colors=[C_DARK, C_DARK, C_DARK, C_GREEN, C_BLUE,
                    taux_color(c['tauxRetrait']),
                    C_GREEN if ecart >= 0 else C_RED, C_DARK],
            bold_cols=[2, 6])

    r = len(coords) + 4
    totals_row(ws, r, 8, ['TOTAL','',
        fmt(sum(c['total']    for c in coords)),
        fmt(sum(c['retires']  for c in coords)),
        fmt(sum(c['restants'] for c in coords)),
        fmtT(data['tauxRetrait']), '', ''])

    row = r + 2
    row = section_title(ws, row, "ANALYSE PAR COORDINATION", 8)
    analyse_block(ws, row, analyse_direction(data), 8)

    if coords:
        chart = BarChart()
        chart.type = "col"; chart.title = "Taux de retrait par coordination"
        chart.y_axis.title = "Taux (%)"; chart.height = 12; chart.width = 22; chart.style = 10
        end = len(coords) + 3
        chart.add_data(Reference(ws, min_col=6, min_row=3, max_row=end), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=2, min_row=4, max_row=end))
        ws.add_chart(chart, f"A{row + len(analyse_direction(data)) + 2}")

# ─── ONGLET AGENCES ──────────────────────────────────────────────────────────

def create_agences_sheet(wb, data, ctx, niveau):
    ws     = wb.create_sheet("🏪 Par Agence")
    agences = sorted(data.get('agences', []), key=lambda a: a['taux_retrait'], reverse=True)
    bg     = NIVEAU_COLOR[niveau]

    if niveau == 'direction':
        sous    = "Toutes agences — Vue nationale"
        headers = ['Rang','Agence','Coordination','Sites','Agents','Total','Retirées','Taux']
        widths  = [7, 26, 22, 8, 8, 14, 14, 12]
    else:
        sous    = f"Agences de la coordination {ctx.get('coordination','')}"
        headers = ['Rang','Agence','Sites','Agents','Total','Retirées','Restantes','Taux','Contribution retard']
        widths  = [7, 30, 8, 8, 14, 14, 14, 12, 20]

    sheet_title(ws, "STATISTIQUES PAR AGENCE", sous, len(headers), bg)
    header_row(ws, 3, headers, widths, bg)

    total_rest = sum(a['cartes_restantes'] for a in agences)

    for i, ag in enumerate(agences):
        r = i + 4
        t = ag['taux_retrait']
        if niveau == 'direction':
            data_row(ws, r,
                [i+1, ag['agence_nom'], ag['coordination_nom'], ag['nombre_sites'],
                 ag['nombre_agents'], fmt(ag['total_cartes']), fmt(ag['cartes_retirees']), fmtT(t)],
                aligns=['center','left','left','center','center','center','center','center'],
                colors=[C_DARK]*6+[C_GREEN, taux_color(t)],
                bold_cols=[2, 8])
        else:
            contrib = (ag['cartes_restantes'] / total_rest * 100) if total_rest else 0
            data_row(ws, r,
                [i+1, ag['agence_nom'], ag['nombre_sites'], ag['nombre_agents'],
                 fmt(ag['total_cartes']), fmt(ag['cartes_retirees']), fmt(ag['cartes_restantes']),
                 fmtT(t), f"{contrib:.0f}% du retard coord."],
                aligns=['center','left','center','center','center','center','center','center','center'],
                colors=[C_DARK,C_DARK,C_DARK,C_DARK,C_DARK,C_GREEN,C_BLUE,taux_color(t),
                        C_RED if contrib>30 else C_ORANGE if contrib>15 else C_DARK],
                bold_cols=[2, 8])

    r = len(agences) + 4
    if niveau == 'direction':
        totals_row(ws, r, len(headers), [
            'TOTAL','','','','',
            fmt(sum(a['total_cartes']    for a in agences)),
            fmt(sum(a['cartes_retirees'] for a in agences)), ''])
    else:
        totals_row(ws, r, len(headers), [
            'TOTAL','','','',
            fmt(sum(a['total_cartes']    for a in agences)),
            fmt(sum(a['cartes_retirees'] for a in agences)),
            fmt(total_rest), '', ''])

    row = r + 2
    row = section_title(ws, row, "ANALYSE", len(headers))
    lignes = analyse_direction(data) if niveau=='direction' else analyse_coordination(data, ctx)
    analyse_block(ws, row, lignes, len(headers))

# ─── ONGLET SITES ─────────────────────────────────────────────────────────────

def create_sites_sheet(wb, data, ctx, niveau):
    ws    = wb.create_sheet("📍 Par Site")
    sites = sorted(data.get('sites', []), key=lambda s: s['tauxRetrait'], reverse=True)
    bg    = NIVEAU_COLOR[niveau]

    if niveau == 'direction':
        sous    = "Tous sites — Vue nationale"
        headers = ['Rang','Site','Coordination','Total','Retirées','Restantes','Taux','Statut']
        widths  = [7, 30, 22, 12, 12, 12, 10, 18]
    elif niveau == 'coordination':
        sous    = f"Sites de la coordination {ctx.get('coordination','')}"
        headers = ['Rang','Site','Agence','Total','Retirées','Restantes','Taux','Statut']
        widths  = [7, 30, 22, 12, 12, 12, 10, 18]
    elif niveau == 'agence':
        sous    = f"Sites de l'agence {ctx.get('agence','')}"
        headers = ['Rang','Site','Total','Retirées','Restantes','Taux','Statut','Priorité']
        widths  = [7, 34, 14, 14, 14, 10, 18, 16]
    else:
        sous    = f"Site : {ctx.get('site','')}"
        headers = ['Indicateur','Valeur','Détail']
        widths  = [30, 20, 40]

    sheet_title(ws, "STATISTIQUES PAR SITE", sous, len(headers), bg)

    # Vue opérateur : tableau vertical
    if niveau == 'site':
        header_row(ws, 3, headers, widths, bg)
        rows_data = [
            ("Total cartes affectées", fmt(data.get('total',0)),   "Nombre total de cartes à distribuer sur ce site"),
            ("Cartes retirées",        fmt(data.get('retires',0)),  "Requérants ayant retiré leur carte"),
            ("Cartes restantes",       fmt(data.get('restants',0)), "Requérants n'ayant pas encore retiré"),
            ("Taux de retrait",        fmtT(data.get('tauxRetrait',0)), taux_label(data.get('tauxRetrait',0))),
        ]
        for i, (ind, val, det) in enumerate(rows_data):
            r  = i + 4
            bg_row = C_LIGHT if i % 2 else C_WHITE
            wcell(ws, r, 1, ind, bold=True, bg=bg_row)
            wcell(ws, r, 2, val, bold=True, align='center',
                  color=taux_color(data.get('tauxRetrait',0)) if 'Taux' in ind else C_DARK, bg=bg_row)
            wcell(ws, r, 3, det, color=C_GRAY, bg=bg_row)
            ws.row_dimensions[r].height = 20
        row = 9
        row = section_title(ws, row, "POSITION ET ANALYSE", len(headers))
        analyse_block(ws, row, analyse_site(data, ctx), len(headers))
        return

    header_row(ws, 3, headers, widths, bg)

    for i, s in enumerate(sites):
        r = i + 4
        t = s['tauxRetrait']
        if niveau in ('direction','coordination'):
            grouper = s.get('coordination','') if niveau=='direction' else s.get('agence', s.get('coordination',''))
            data_row(ws, r,
                [i+1, s['site'], grouper, fmt(s['total']),
                 fmt(s['retires']), fmt(s['restants']), fmtT(t), taux_label(t)],
                aligns=['center','left','left','center','center','center','center','left'],
                colors=[C_DARK,C_DARK,C_GRAY,C_DARK,C_GREEN,C_BLUE,taux_color(t),C_DARK],
                bold_cols=[2, 7])
        else:
            prio = ('🔴 Urgent' if t < 30 else '🟠 Prioritaire' if t < 50 else '🟡 À surveiller' if t < 75 else '🟢 OK')
            data_row(ws, r,
                [i+1, s['site'], fmt(s['total']), fmt(s['retires']),
                 fmt(s['restants']), fmtT(t), taux_label(t), prio],
                aligns=['center','left','center','center','center','center','left','center'],
                colors=[C_DARK,C_DARK,C_DARK,C_GREEN,C_BLUE,taux_color(t),C_DARK,C_DARK],
                bold_cols=[2, 6])

    r = len(sites) + 4
    pad = [''] * (len(headers) - 6)
    totals_row(ws, r, len(headers), ['TOTAL',''] + pad + [
        fmt(sum(s['total']    for s in sites)),
        fmt(sum(s['retires']  for s in sites)),
        fmt(sum(s['restants'] for s in sites)), '', ''])

    row = r + 2
    row = section_title(ws, row, "ANALYSE DES SITES", len(headers))
    lignes = (analyse_direction(data)       if niveau=='direction'
         else analyse_coordination(data,ctx) if niveau=='coordination'
         else analyse_agence(data, ctx))
    analyse_block(ws, row, lignes, len(headers))

# ─── ONGLET RECOMMANDATIONS ──────────────────────────────────────────────────

def create_recommandations_sheet(wb, data, ctx, niveau):
    ws    = wb.create_sheet("💡 Recommandations")
    bg    = NIVEAU_COLOR[niveau]
    sites = data.get('sites', [])
    t     = data.get('tauxRetrait', 0)
    seuil = {'direction':60,'coordination':55,'agence':50,'site':50}[niveau]

    titres = {
        'direction':    "RECOMMANDATIONS STRATÉGIQUES — DIRECTION",
        'coordination': "RECOMMANDATIONS — COORDINATION",
        'agence':       "RECOMMANDATIONS OPÉRATIONNELLES — AGENCE",
        'site':         "ACTIONS À MENER — SITE",
    }
    sheet_title(ws, titres[niveau], f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y')}", 6, bg)

    alertes = sorted([s for s in sites if s['tauxRetrait'] < seuil], key=lambda s: s['tauxRetrait'])
    row = 3
    row = section_title(ws, row, f"⚠️ ALERTES — Sites avec taux < {seuil}% ({len(alertes)} site(s))", 6, C_FAIL)

    if alertes:
        header_row(ws, row, ['Site','Coord./Agence','Total','Retirées','Restantes','Taux'],
                   [30,22,12,12,12,12], C_RED)
        row += 1
        for i, s in enumerate(alertes):
            grouper = s.get('coordination','') if niveau in ('direction','coordination') else s.get('agence','')
            data_row(ws, row,
                [s['site'], grouper, fmt(s['total']), fmt(s['retires']), fmt(s['restants']), fmtT(s['tauxRetrait'])],
                aligns=['left','left','center','center','center','center'],
                colors=[C_DARK]*5+[C_RED], bold_cols=[6])
            row += 1
    else:
        ws.merge_cells(f'A{row}:F{row}')
        wcell(ws, row, 1, f"✅ Aucun site sous le seuil de {seuil}%.", color=C_GREEN, bg=C_SUCC)
        ws.row_dimensions[row].height = 22; row += 1

    row += 1
    top10 = sorted(sites, key=lambda s: s['tauxRetrait'], reverse=True)[:10]
    row = section_title(ws, row, "🏆 TOP SITES PERFORMANTS", 6, C_SUCC)
    header_row(ws, row, ['Rang','Site','Coord./Agence','Total','Retirées','Taux'],
               [7,30,22,12,12,12], C_GREEN)
    row += 1
    for i, s in enumerate(top10):
        grouper = s.get('coordination','') if niveau in ('direction','coordination') else s.get('agence','')
        data_row(ws, row,
            [i+1, s['site'], grouper, fmt(s['total']), fmt(s['retires']), fmtT(s['tauxRetrait'])],
            aligns=['center','left','left','center','center','center'],
            colors=[C_DARK]*5+[C_GREEN], bold_cols=[6])
        row += 1

    row += 1
    plans = {
        'direction': {
            'titre': "PLAN D'ACTION STRATÉGIQUE",
            75: ["✅ Performance nationale excellente.",
                 "→ Documenter les processus des coordinations leaders et créer un guide de bonnes pratiques.",
                 "→ Organiser des échanges entre meilleures coordinations et celles en retard.",
                 "→ Engager la phase de clôture pour les requérants n'ayant pas retiré leur carte.",
                 "→ Préparer le rapport final de distribution pour les autorités compétentes."],
            50: ["📈 Performance nationale satisfaisante — hétérogénéité à réduire.",
                 f"→ Déclencher un audit sur les {len(alertes)} coordination(s) sous le seuil de 60%.",
                 "→ Réallouer des ressources des zones performantes vers les zones critiques.",
                 "→ Instaurer un reporting hebdomadaire national avec indicateurs par coordination.",
                 "→ Identifier les blocages opérationnels empêchant les requérants de retirer leur carte."],
            0:  ["🚨 Situation critique — mobilisation nationale immédiate.",
                 f"→ Constituer une cellule de crise nationale avec réunion quotidienne.",
                 "→ Déployer des équipes renforcées dans toutes les coordinations sous 50%.",
                 "→ Lancer une campagne de sensibilisation multicanal pour les requérants non venus.",
                 "→ Fixer des objectifs hebdomadaires contraignants avec reporting quotidien.",
                 "→ Envisager des points de distribution mobiles dans les zones isolées."],
        },
        'coordination': {
            'titre': "PLAN D'ACTION POUR LA COORDINATION",
            75: ["✅ Votre coordination affiche une performance excellente.",
                 "→ Partager les bonnes pratiques des meilleures agences avec les agences en retard.",
                 "→ Finaliser les relances auprès des requérants restants.",
                 "→ Maintenir le suivi hebdomadaire des sites."],
            50: ["📈 Performance satisfaisante — efforts ciblés nécessaires.",
                 f"→ Demander un rapport hebdomadaire aux chefs d'équipe des agences en retard.",
                 "→ Fixer un objectif de taux cible par agence pour la prochaine période.",
                 "→ Vérifier que les requérants des sites critiques ont bien reçu une notification.",
                 "→ Envisager des permanences supplémentaires sur les sites à fort volume restant."],
            0:  ["🚨 Performance insuffisante dans la coordination — action urgente.",
                 f"→ Mobilisation immédiate sur les {len(alertes)} site(s) en alerte.",
                 "→ Renforcer les équipes terrain et étendre les horaires d'ouverture.",
                 "→ Lancer une campagne de relance ciblée auprès des requérants non venus.",
                 "→ Instaurer un suivi quotidien avec point de situation chaque matin."],
        },
        'agence': {
            'titre': "PLAN D'ACTION OPÉRATIONNEL — AGENCE",
            75: ["✅ Votre agence est en bonne performance.",
                 "→ Accompagner les sites encore en dessous de 75%.",
                 "→ Signaler les requérants injoignables ou avec des problèmes de documents à la coordination."],
            50: ["📈 Performance correcte — focus sur les sites en retard.",
                 f"→ Intervention prioritaire sur les {len(alertes)} site(s) sous 50% cette semaine.",
                 "→ Vérifier que tous les requérants des sites critiques ont été contactés.",
                 "→ Envisager une permanence supplémentaire sur les sites à fort volume restant.",
                 "→ Remonter les blocages identifiés à la coordination."],
            0:  ["🚨 Performance insuffisante — action immédiate requise.",
                 f"→ {len(alertes)} site(s) en alerte critique nécessitent une intervention sous 48h.",
                 "→ Déployer un agent mobile sur les sites les plus en retard.",
                 "→ Contacter personnellement les responsables des sites en alerte.",
                 "→ Signaler la situation à la coordination et demander des renforts.",
                 "→ Mettre en place un point quotidien avec les responsables de sites."],
        },
        'site': {
            'titre': "ACTIONS IMMÉDIATES POUR LE SITE",
            75: ["✅ Votre site est en bonne performance.",
                 "→ Maintenir le rythme actuel de distribution.",
                 "→ Signaler à votre chef d'équipe les requérants injoignables.",
                 "→ Vérifier que les requérants restants ont bien reçu une notification."],
            50: ["💡 Des requérants n'ont pas encore retiré leur carte.",
                 "→ Établir la liste des requérants n'ayant pas retiré et tenter un contact.",
                 "→ Signaler à votre chef d'équipe les cas de requérants injoignables.",
                 "→ Vérifier que les horaires d'ouverture sont bien communiqués.",
                 "→ Orienter les requérants ayant des problèmes de documents vers le bon interlocuteur."],
            0:  ["🚨 Taux critique — intervention requise.",
                 "→ Établir immédiatement la liste complète des requérants n'ayant pas retiré leur carte.",
                 "→ Lancer les relances téléphoniques ou par affichage local.",
                 "→ Informer votre chef d'équipe et demander un renfort ou une permanence supplémentaire.",
                 "→ Identifier les requérants avec problèmes de documents et les orienter.",
                 "→ Tenir un registre quotidien des retraits jusqu'à atteindre l'objectif."],
        },
    }

    plan   = plans[niveau]
    seuil_p = 75 if t >= 75 else 50 if t >= 50 else 0
    row = section_title(ws, row, f"📋 {plan['titre']}", 6, C_WARN)
    for action in plan[seuil_p]:
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row, 1, action)
        if any(action.startswith(p) for p in ['✅','📈','🚨','💡']):
            c.font = _font(True, C_DARK, 10);  c.fill = _fill(C_WARN)
        else:
            c.font = _font(False, C_DARK, 10); c.fill = _fill(C_WHITE)
        c.alignment = _align('left', True)
        ws.row_dimensions[row].height = 22
        row += 1

    for i in range(1, 7):
        cw(ws, i, 20)

# ─── POINT D'ENTRÉE ──────────────────────────────────────────────────────────

def generate(data_str):
    data   = json.loads(data_str)
    ctx    = data.get('_contexte', {})
    niveau = ctx.get('niveau', 'direction')

    wb = Workbook()
    create_resume_sheet(wb, data, ctx, niveau)

    if niveau == 'direction':
        create_coordinations_sheet(wb, data, ctx)
        create_agences_sheet(wb, data, ctx, niveau)
        create_sites_sheet(wb, data, ctx, niveau)
        create_recommandations_sheet(wb, data, ctx, niveau)

    elif niveau == 'coordination':
        create_agences_sheet(wb, data, ctx, niveau)
        create_sites_sheet(wb, data, ctx, niveau)
        create_recommandations_sheet(wb, data, ctx, niveau)

    elif niveau == 'agence':
        create_sites_sheet(wb, data, ctx, niveau)
        create_recommandations_sheet(wb, data, ctx, niveau)

    elif niveau == 'site':
        create_recommandations_sheet(wb, data, ctx, niveau)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()

if __name__ == '__main__':
    print(generate(sys.argv[1]))